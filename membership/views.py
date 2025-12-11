from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login, logout, authenticate
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.http import HttpResponse
from django.utils import timezone
from .models import Member, Child, MembershipFee, Payment, UserProfile
from .forms import (
    LoginForm, RegisterForm, MemberForm, ChildFormSet, 
    MembershipFeeForm, PaymentForm
)
from datetime import datetime, timedelta

import pandas as pd
import openpyxl
from io import BytesIO
from django.core.exceptions import ValidationError
from django.db import transaction



# Helper function to check if user is admin/staff
def is_admin(user):
    return user.is_superuser or user.is_staff


# Authentication Views
def user_login(request):
    """User login view"""
    if request.user.is_authenticated:
        return redirect('membership:home')
    
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Welcome back, {user.first_name or user.username}!')
                return redirect('membership:home')
        else:
            messages.error(request, 'Invalid username or password.')
    else:
        form = LoginForm()
    
    return render(request, 'membership/login.html', {'form': form})


def user_register(request):
    """User registration view"""
    if request.user.is_authenticated:
        return redirect('membership:home')
    
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Registration successful! Your account is pending admin approval.')
            return redirect('membership:pending_approval')
        else:
            messages.error(request, 'Registration failed. Please check the form.')
    else:
        form = RegisterForm()
    
    return render(request, 'membership/register.html', {'form': form})


def user_logout(request):
    """User logout view"""
    logout(request)
    messages.info(request, 'You have been logged out.')
    return redirect('membership:login')


@login_required
def pending_approval(request):
    """View for users waiting for admin approval"""
    return render(request, 'membership/pending_approval.html')


@login_required
def home(request):
    """Enhanced dashboard with statistics, upcoming birthdays, and insights"""
    
    # Calculate statistics
    total_members = Member.objects.filter(is_active=True).count()
    regular_members = Member.objects.filter(
        is_active=True, 
        membership_type='REGULAR'
    ).count()
    lifetime_members = Member.objects.filter(
        is_active=True, 
        membership_type='LIFETIME'
    ).count()
    honarary_members = Member.objects.filter(
        is_active=True, 
        membership_type='HONARARY'
    ).count()
    
    total_revenue = Payment.objects.aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    # Calculate percentages
    if total_members > 0:
        regular_percentage = round((regular_members / total_members * 100), 1)
        lifetime_percentage = round((lifetime_members / total_members * 100), 1)
        honarary_percentage = round((honarary_members / total_members * 100), 1)
    else:
        regular_percentage = 0
        lifetime_percentage = 0
        honarary_percentage = 0
    
    # Get upcoming birthdays (next 7 days)
    today = timezone.now().date()
    upcoming_birthdays = []
    
    # Get all members with birthdates
    members = Member.objects.filter(
        is_active=True,
        date_of_birth__isnull=False
    )
    
    for member in members:
        # Calculate this year's birthday
        try:
            birthday_this_year = member.date_of_birth.replace(year=today.year)
        except ValueError:
            # Handle leap year babies (Feb 29)
            birthday_this_year = member.date_of_birth.replace(year=today.year, day=28)
        
        # If birthday already passed this year, use next year
        if birthday_this_year < today:
            try:
                birthday_this_year = member.date_of_birth.replace(year=today.year + 1)
            except ValueError:
                birthday_this_year = member.date_of_birth.replace(year=today.year + 1, day=28)
        
        # Calculate days until birthday
        days_until = (birthday_this_year - today).days
        
        # Include if within next 7 days
        if 0 <= days_until <= 7:
            member.days_until_birthday = days_until
            upcoming_birthdays.append(member)
    
    # Sort by days until birthday
    upcoming_birthdays.sort(key=lambda x: x.days_until_birthday)
    
    # Get recent payments (last 5)
    recent_payments = Payment.objects.select_related(
        'member', 'membership_fee'
    ).order_by('-payment_date')[:5]
    
    # Get memberships expiring in next 30 days
    thirty_days_from_now = today + timedelta(days=30)
    expiring_soon = Member.objects.filter(
        is_active=True,
        membership_type='REGULAR',
        membership_valid_until__isnull=False,
        membership_valid_until__lte=thirty_days_from_now,
        membership_valid_until__gte=today
    ).order_by('membership_valid_until')[:5]
    
    # Add days_remaining to each expiring member
    for member in expiring_soon:
        if member.membership_valid_until:
            member.days_remaining = (member.membership_valid_until - today).days
        else:
            member.days_remaining = 0
    
    context = {
        'stats': {
            'total_members': total_members,
            'active_members': total_members,
            'regular_members': regular_members,
            'lifetime_members': lifetime_members,
            'honarary_members': honarary_members,
            'total_revenue': total_revenue,
            'regular_percentage': regular_percentage,
            'lifetime_percentage': lifetime_percentage,
            'honarary_percentage': honarary_percentage,
        },
        'upcoming_birthdays': upcoming_birthdays[:10],  # Show max 10
        'recent_payments': recent_payments,
        'expiring_soon': expiring_soon,
    }
    
    return render(request, 'membership/home.html', context)

@login_required
def member_list(request):
    """List all members with search and filter"""
    members = Member.objects.all()
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        members = members.filter(
            Q(name__icontains=search_query) |
            Q(membership_number__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    # Filter by membership type
    membership_type = request.GET.get('type', '')
    if membership_type:
        members = members.filter(membership_type=membership_type)
    
    # Filter by status
    status = request.GET.get('status', '')
    if status:
        is_active = status == 'active'
        members = members.filter(is_active=is_active)
    
    context = {
        'members': members.order_by('-join_date'),
        'search_query': search_query,
        'membership_type': membership_type,
        'status': status,
    }
    return render(request, 'membership/member_list.html', context)


@login_required
def member_detail(request, pk):
    member = get_object_or_404(Member, pk=pk)
    children = member.children.all()
    payments = Payment.objects.filter(member=member).order_by('-payment_date')
    
    # Calculate membership expiry status
    from django.utils import timezone
    today = timezone.now().date()
    
    expiry_status = {
        'has_paid': payments.exists(),  # Has made any payment
        'is_lifetime': member.membership_type == 'LIFETIME',
        'has_expiry': False,
        'is_expired': False,
        'is_expiring_soon': False,
        'days_remaining': None,
        'days_overdue': None,
        'status_class': '',  # CSS class for badge
        'status_text': '',   # Display text
        'status_icon': '',   # Bootstrap icon
    }
    
    # Determine expiry status
    if not expiry_status['has_paid']:
        expiry_status['status_class'] = 'bg-secondary'
        expiry_status['status_text'] = 'No Payment Made'
        expiry_status['status_icon'] = 'bi-x-circle'
    elif expiry_status['is_lifetime']:
        expiry_status['status_class'] = 'bg-success'
        expiry_status['status_text'] = 'Lifetime Membership'
        expiry_status['status_icon'] = 'bi-infinity'
    else:
        # Regular membership - check expiry
        expiry_status['has_expiry'] = True
        
        if member.membership_valid_until:
            if member.membership_valid_until < today:
                # Expired
                expiry_status['is_expired'] = True
                expiry_status['days_overdue'] = (today - member.membership_valid_until).days
                expiry_status['status_class'] = 'bg-danger'
                expiry_status['status_text'] = f'Expired ({expiry_status["days_overdue"]} days ago)'
                expiry_status['status_icon'] = 'bi-exclamation-triangle-fill'
            else:
                # Active - check if expiring soon
                expiry_status['days_remaining'] = (member.membership_valid_until - today).days
                
                if expiry_status['days_remaining'] <= 7:
                    expiry_status['is_expiring_soon'] = True
                    expiry_status['status_class'] = 'bg-danger'
                    expiry_status['status_text'] = f'Expiring Soon ({expiry_status["days_remaining"]} days left)'
                    expiry_status['status_icon'] = 'bi-exclamation-circle-fill'
                elif expiry_status['days_remaining'] <= 30:
                    expiry_status['is_expiring_soon'] = True
                    expiry_status['status_class'] = 'bg-warning'
                    expiry_status['status_text'] = f'Expiring in {expiry_status["days_remaining"]} days'
                    expiry_status['status_icon'] = 'bi-clock-fill'
                else:
                    expiry_status['status_class'] = 'bg-success'
                    expiry_status['status_text'] = f'Active ({expiry_status["days_remaining"]} days remaining)'
                    expiry_status['status_icon'] = 'bi-check-circle-fill'
        else:
            # No expiry date set
            expiry_status['status_class'] = 'bg-warning'
            expiry_status['status_text'] = 'No Expiry Date Set'
            expiry_status['status_icon'] = 'bi-calendar-x'
    
    context = {
        'member': member,
        'children': children,
        'payments': payments,
        'total_paid' : member.get_total_paid(),
        'expiry_status': expiry_status,  # Add this
    }
    
    return render(request, 'membership/member_detail.html', context)


@login_required
def member_add(request):
    """Add new member"""
    if request.method == 'POST':
        form = MemberForm(request.POST, request.FILES)
        formset = ChildFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            member = form.save()
            formset.instance = member
            formset.save()
            
            messages.success(request, f'Member {member.name} added successfully! Membership Number: {member.membership_number}')
            return redirect('membership:member_detail', pk=member.pk)
    else:
        form = MemberForm()
        formset = ChildFormSet()
    
    context = {
        'form': form,
        'formset': formset,
        'action': 'Add',
    }
    return render(request, 'membership/member_form.html', context)


@login_required
def member_edit(request, pk):
    """Edit existing member"""
    member = get_object_or_404(Member, pk=pk)
    
    if request.method == 'POST':
        form = MemberForm(request.POST, request.FILES, instance=member)
        formset = ChildFormSet(request.POST, instance=member)
        
        if form.is_valid() and formset.is_valid():
            member = form.save()
            formset.save()
            
            messages.success(request, f'Member {member.name} updated successfully!')
            return redirect('membership:member_detail', pk=member.pk)
    else:
        form = MemberForm(instance=member)
        formset = ChildFormSet(instance=member)
    
    context = {
        'form': form,
        'formset': formset,
        'member': member,
        'action': 'Edit',
    }
    return render(request, 'membership/member_form.html', context)


@login_required
def member_delete(request, pk):
    """Delete member (permanent deletion)"""
    member = get_object_or_404(Member, pk=pk)
    
    if request.method == 'POST':
        member_name = member.name
        member.delete()
        messages.success(request, f'Member {member_name} has been permanently deleted.')
        return redirect('membership:member_list')
    
    context = {'member': member}
    return render(request, 'membership/member_confirm_delete.html', context)


@login_required
def payment_list(request):
    """List all payments with filters"""
    payments = Payment.objects.all()
    
    # Filter by date range
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    if start_date:
        payments = payments.filter(payment_date__gte=start_date)
    if end_date:
        payments = payments.filter(payment_date__lte=end_date)
    
    # Filter by payment mode
    payment_mode = request.GET.get('payment_mode', '')
    if payment_mode:
        payments = payments.filter(payment_mode=payment_mode)
    
    # Calculate totals
    total_amount = payments.aggregate(total=Sum('amount'))['total'] or 0
    
    context = {
        'payments': payments.order_by('-payment_date'),
        'total_amount': total_amount,
        'start_date': start_date,
        'end_date': end_date,
        'payment_mode': payment_mode,
    }
    return render(request, 'membership/payment_list.html', context)


@login_required
def payment_add(request):
    """Add new payment"""
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        
        if form.is_valid():
            payment = form.save()
            messages.success(request, f'Payment recorded successfully! Receipt Number: {payment.receipt_number}')
            return redirect('membership:payment_receipt', pk=payment.pk)
    else:
        form = PaymentForm()
        # Pre-fill collected_by with current user's name
        form.initial['collected_by'] = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username
    
    context = {'form': form}
    return render(request, 'membership/payment_form.html', context)


@login_required
def payment_edit(request, pk):
    """Edit existing payment"""
    payment = get_object_or_404(Payment, pk=pk)
    
    if request.method == 'POST':
        form = PaymentForm(request.POST, instance=payment)
        
        if form.is_valid():
            payment = form.save()
            messages.success(request, f'Payment {payment.receipt_number} updated successfully!')
            return redirect('membership:payment_receipt', pk=payment.pk)
    else:
        form = PaymentForm(instance=payment)
    
    context = {
        'form': form,
        'payment': payment,
    }
    return render(request, 'membership/payment_form.html', context)


@login_required
def payment_delete(request, pk):
    """Delete payment"""
    payment = get_object_or_404(Payment, pk=pk)
    
    if request.method == 'POST':
        receipt_number = payment.receipt_number
        payment.delete()
        messages.success(request, f'Payment {receipt_number} has been deleted.')
        return redirect('membership:payment_list')
    
    context = {'payment': payment}
    return render(request, 'membership/payment_confirm_delete.html', context)


@login_required
def payment_receipt(request, pk):
    """Generate and display payment receipt"""
    payment = get_object_or_404(Payment, pk=pk)
    
    context = {
        'payment': payment,
        'print_date': timezone.now(),
    }
    return render(request, 'membership/payment_receipt.html', context)


@login_required
def fee_list(request):
    """List all membership fees"""
    fees = MembershipFee.objects.all().order_by('membership_type', 'payment_frequency')
    
    context = {'fees': fees}
    return render(request, 'membership/fee_list.html', context)


@login_required
def fee_add(request):
    """Add new membership fee"""
    if request.method == 'POST':
        form = MembershipFeeForm(request.POST)
        
        if form.is_valid():
            fee = form.save()
            messages.success(request, 'Membership fee structure added successfully!')
            return redirect('membership:fee_list')
    else:
        form = MembershipFeeForm()
    
    context = {'form': form}
    return render(request, 'membership/fee_form.html', context)


@login_required
def fee_edit(request, pk):
    """Edit existing membership fee"""
    fee = get_object_or_404(MembershipFee, pk=pk)
    
    if request.method == 'POST':
        form = MembershipFeeForm(request.POST, instance=fee)
        
        if form.is_valid():
            fee = form.save()
            messages.success(request, 'Membership fee structure updated successfully!')
            return redirect('membership:fee_list')
    else:
        form = MembershipFeeForm(instance=fee)
    
    context = {
        'form': form,
        'fee': fee,
    }
    return render(request, 'membership/fee_form.html', context)


@login_required
def fee_delete(request, pk):
    """Delete membership fee"""
    fee = get_object_or_404(MembershipFee, pk=pk)
    
    if request.method == 'POST':
        fee.delete()
        messages.success(request, 'Membership fee structure has been deleted.')
        return redirect('membership:fee_list')
    
    context = {'fee': fee}
    return render(request, 'membership/fee_confirm_delete.html', context)


@login_required
def revenue_report(request):
    """Display revenue collection reports"""
    # Get date range from request or default to current month
    today = timezone.now().date()
    start_date = request.GET.get('start_date', today.replace(day=1).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    
    # Filter payments by date range
    payments = Payment.objects.filter(
        payment_date__range=[start_date, end_date]
    )
    
    # Calculate statistics
    total_revenue = payments.aggregate(total=Sum('amount'))['total'] or 0
    payment_count = payments.count()
    
    # Revenue by payment mode
    revenue_by_mode = payments.values('payment_mode').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # Revenue by membership type
    revenue_by_type = payments.values('membership_fee__membership_type').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # Daily revenue (for chart)
    daily_revenue = payments.extra(
        select={'day': 'date(payment_date)'}
    ).values('day').annotate(
        total=Sum('amount')
    ).order_by('day')
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_revenue': total_revenue,
        'payment_count': payment_count,
        'revenue_by_mode': revenue_by_mode,
        'revenue_by_type': revenue_by_type,
        'daily_revenue': daily_revenue,
        'payments': payments.order_by('-payment_date'),
    }
    return render(request, 'membership/revenue_report.html', context)


# User Approval Views (Admin Only)
@login_required
@user_passes_test(is_admin)
def user_approval_list(request):
    """List all users for approval management (admin only)"""
    filter_type = request.GET.get('filter', 'pending')
    
    # Get all user profiles
    if filter_type == 'pending':
        users = UserProfile.objects.filter(is_approved=False).select_related('user', 'approved_by')
    elif filter_type == 'approved':
        users = UserProfile.objects.filter(is_approved=True).select_related('user', 'approved_by')
    else:  # all
        users = UserProfile.objects.all().select_related('user', 'approved_by')
    
    # Order by date joined (newest first)
    users = users.order_by('-user__date_joined')
    
    # Handle bulk actions
    if request.method == 'POST':
        action = request.POST.get('action')
        user_ids = request.POST.getlist('user_ids')
        
        if user_ids:
            profiles = UserProfile.objects.filter(id__in=user_ids)
            
            if action == 'bulk_approve':
                count = 0
                for profile in profiles:
                    if not profile.is_approved:
                        profile.is_approved = True
                        profile.approved_by = request.user
                        profile.approved_at = timezone.now()
                        profile.save()
                        count += 1
                messages.success(request, f'{count} user(s) approved successfully.')
            
            elif action == 'bulk_unapprove':
                count = profiles.update(is_approved=False, approved_by=None, approved_at=None)
                messages.warning(request, f'{count} user(s) unapproved.')
        
        return redirect('membership:user_approval_list')
    
    # Get counts
    pending_count = UserProfile.objects.filter(is_approved=False).count()
    approved_count = UserProfile.objects.filter(is_approved=True).count()
    all_count = UserProfile.objects.count()
    
    context = {
        'users': users,
        'filter': filter_type,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'all_count': all_count,
    }
    
    return render(request, 'membership/user_approval_list.html', context)


@login_required
@user_passes_test(is_admin)
def user_detail_admin(request, pk):
    """View user details for admin"""
    user_profile = get_object_or_404(UserProfile, pk=pk)
    
    context = {
        'user_profile': user_profile,
    }
    
    return render(request, 'membership/user_detail_admin.html', context)


@login_required
@user_passes_test(is_admin)
def user_approve(request, pk):
    """Approve a user"""
    user_profile = get_object_or_404(UserProfile, pk=pk)
    
    if not user_profile.is_approved:
        user_profile.is_approved = True
        user_profile.approved_by = request.user
        user_profile.approved_at = timezone.now()
        user_profile.save()
        messages.success(request, f'User {user_profile.user.username} has been approved.')
    else:
        messages.info(request, f'User {user_profile.user.username} is already approved.')
    
    return redirect('membership:user_approval_list')


@login_required
@user_passes_test(is_admin)
def user_unapprove(request, pk):
    """Unapprove a user"""
    user_profile = get_object_or_404(UserProfile, pk=pk)
    
    if user_profile.is_approved:
        user_profile.is_approved = False
        user_profile.approved_by = None
        user_profile.approved_at = None
        user_profile.save()
        messages.warning(request, f'User {user_profile.user.username} has been unapproved.')
    else:
        messages.info(request, f'User {user_profile.user.username} is already unapproved.')
    
    return redirect('membership:user_approval_list')



from django.utils import timezone
from datetime import timedelta
from django.db.models import Q, F, ExpressionWrapper, fields

@login_required
def renewal_required_report(request):
    """Members requiring renewal - expired or expiring soon"""
    today = timezone.now().date()
    thirty_days = today + timedelta(days=30)
    
    # Expired members
    expired_members = Member.objects.filter(
        membership_type='REGULAR',
        is_active=True,
        membership_valid_until__lt=today
    )
    
    # Expiring soon (next 30 days)
    expiring_soon = Member.objects.filter(
        membership_type='REGULAR',
        is_active=True,
        membership_valid_until__gte=today,
        membership_valid_until__lte=thirty_days
    )
    
    context = {
        'expired_members': expired_members,
        'expiring_soon_members': expiring_soon,
        'expired_count': expired_members.count(),
        'expiring_soon_count': expiring_soon.count(),
        'total_requiring_renewal': expired_members.count() + expiring_soon.count(),
    }
    
    return render(request, 'membership/renewal_required_report.html', context)

# 2. MEMBERSHIP EXPIRY REPORT - FIXED
@login_required
def membership_expiry_report(request):
    """All memberships with expiry tracking"""
    today = timezone.now().date()
    
    # Get filter parameters
    status = request.GET.get('status', '')
    membership_type_filter = request.GET.get('type', '')
    
    # Base queryset - only regular members have expiry
    members = Member.objects.filter(
        membership_type='REGULAR',
        is_active=True
    ).exclude(membership_valid_until__isnull=True)
    
    # Apply membership type filter
    if membership_type_filter:
        members = members.filter(membership_type=membership_type_filter)
    
    # Apply status filter
    if status == 'expired':
        members = members.filter(membership_valid_until__lt=today)
    elif status == 'this_month':
        end_of_month = today.replace(day=1) + timedelta(days=32)
        end_of_month = end_of_month.replace(day=1) - timedelta(days=1)
        members = members.filter(
            membership_valid_until__gte=today,
            membership_valid_until__lte=end_of_month
        )
    elif status == 'next_month':
        next_month_start = (today.replace(day=1) + timedelta(days=32)).replace(day=1)
        next_month_end = (next_month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        members = members.filter(
            membership_valid_until__gte=next_month_start,
            membership_valid_until__lte=next_month_end
        )
    elif status == 'later':
        next_month_start = (today.replace(day=1) + timedelta(days=32)).replace(day=1)
        next_month_end = (next_month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        members = members.filter(membership_valid_until__gt=next_month_end)
    
    # Convert to list and add calculated fields
    members_list = list(members.order_by('membership_valid_until'))
    
    for member in members_list:
        if member.membership_valid_until:
            if member.membership_valid_until < today:
                member.is_expired_custom = True
                member.days_overdue_custom = (today - member.membership_valid_until).days
            else:
                member.is_expired_custom = False
                member.days_until_expiry_custom = (member.membership_valid_until - today).days
                # Calculate percentage for progress bar (0-100)
                days_passed = 365 - member.days_until_expiry_custom
                member.expiry_percentage = min(100, max(0, (days_passed / 365) * 100))
    
    # Calculate stats
    all_regular = Member.objects.filter(membership_type='REGULAR', is_active=True)
    end_of_month = today.replace(day=1) + timedelta(days=32)
    end_of_month = end_of_month.replace(day=1) - timedelta(days=1)
    next_month_start = (today.replace(day=1) + timedelta(days=32)).replace(day=1)
    next_month_end = (next_month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    
    context = {
        'members': members_list,
        'expired_count': all_regular.filter(membership_valid_until__lt=today).count(),
        'this_month_count': all_regular.filter(
            membership_valid_until__gte=today,
            membership_valid_until__lte=end_of_month
        ).count(),
        'next_month_count': all_regular.filter(
            membership_valid_until__gte=next_month_start,
            membership_valid_until__lte=next_month_end
        ).count(),
        'later_count': all_regular.filter(membership_valid_until__gt=next_month_end).count(),
        'status': status,
        'membership_type': membership_type_filter,
    }
    
    return render(request, 'membership/membership_expiry_report.html', context)

# 3. NEW MEMBERS REPORT
@login_required
def new_members_report(request):
    """Recent member registrations"""
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    membership_type = request.GET.get('type', '')
    
    # Default to last 30 days
    if not end_date:
        end_date = timezone.now().date()
    else:
        end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
    
    if not start_date:
        start_date = end_date - timedelta(days=30)
    else:
        start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
    
    # Get new members
    members = Member.objects.filter(
        join_date__gte=start_date,
        join_date__lte=end_date
    ).order_by('-join_date')
    
    # Apply type filter
    if membership_type:
        members = members.filter(membership_type=membership_type)
    
    # Add days since joining
    today = timezone.now().date()
    for member in members:
        member.days_since_joining = (today - member.join_date).days
    
    # Calculate stats
    this_month_start = today.replace(day=1)
    
    context = {
        'members': members,
        'total_new_members': members.count(),
        'this_month_count': Member.objects.filter(
            join_date__gte=this_month_start
        ).count(),
        'regular_count': members.filter(membership_type='REGULAR').count(),
        'lifetime_count': members.filter(membership_type='LIFETIME').count(),
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'membership_type': membership_type,
    }
    
    return render(request, 'membership/new_members_report.html', context)

@login_required
@user_passes_test(is_admin)
def bulk_upload_members(request):
    """Bulk upload members from Excel/CSV file"""
    
    if request.method == 'POST':
        if 'file' not in request.FILES:
            messages.error(request, 'Please select a file to upload.')
            return redirect('membership:bulk_upload_members')
        
        file = request.FILES['file']
        file_extension = file.name.split('.')[-1].lower()
        
        # Validate file type
        if file_extension not in ['xlsx', 'xls', 'csv']:
            messages.error(request, 'Invalid file format. Please upload Excel (.xlsx, .xls) or CSV file.')
            return redirect('membership:bulk_upload_members')
        
        try:
            # Read the file based on extension
            if file_extension == 'csv':
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
            
            # Process and validate data
            success_count = 0
            error_count = 0
            errors = []
            skipped_count = 0
            auto_generated = []
            
            for index, row in df.iterrows():
                try:
                    row_auto_gen = []
                    
                    # Skip completely empty rows
                    if 'name' not in row or pd.isna(row['name']) or str(row['name']).strip() == '':
                        skipped_count += 1
                        continue
                    
                    name = str(row['name']).strip()
                    
                    # Auto-generate phone if missing
                    phone = None
                    if 'phone' in row and not pd.isna(row['phone']) and str(row['phone']).strip():
                        phone = str(row['phone']).strip()
                        # Check if phone already exists
                        if Member.objects.filter(phone=phone).exists():
                            error_count += 1
                            errors.append(f'Row {index + 2}: Phone {phone} already exists')
                            continue
                    else:
                        # Generate placeholder phone (TEMP + timestamp + row number)
                        import time
                        phone = f"TEMP{int(time.time())}{index}"
                        row_auto_gen.append('phone')
                    
                    # Parse dates
                    date_of_birth = None
                    if 'date_of_birth' in row and not pd.isna(row['date_of_birth']):
                        try:
                            date_of_birth = pd.to_datetime(row['date_of_birth']).date()
                        except:
                            pass
                    
                    join_date = timezone.now().date()
                    if 'join_date' in row and not pd.isna(row['join_date']):
                        try:
                            join_date = pd.to_datetime(row['join_date']).date()
                        except:
                            pass
                    
                    citizenship_issue_date = None
                    if 'citizenship_issue_date' in row and not pd.isna(row['citizenship_issue_date']):
                        try:
                            citizenship_issue_date = pd.to_datetime(row['citizenship_issue_date']).date()
                        except:
                            pass
                    
                    # Membership type
                    membership_type = 'REGULAR'
                    if 'membership_type' in row and not pd.isna(row['membership_type']):
                        membership_type = str(row['membership_type']).upper().strip()
                        if membership_type not in ['REGULAR', 'LIFETIME']:
                            membership_type = 'REGULAR'
                            row_auto_gen.append('membership_type')
                    else:
                        row_auto_gen.append('membership_type')
                    
                    # Gender
                    gender = 'O'
                    if 'gender' in row and not pd.isna(row['gender']):
                        gender_value = str(row['gender']).upper().strip()
                        if gender_value in ['M', 'MALE']:
                            gender = 'M'
                        elif gender_value in ['F', 'FEMALE']:
                            gender = 'F'
                        else:
                            row_auto_gen.append('gender')
                    else:
                        row_auto_gen.append('gender')
                    
                    # Payment frequency
                    payment_frequency = 'MONTHLY'
                    if 'payment_frequency' in row and not pd.isna(row['payment_frequency']):
                        freq_value = str(row['payment_frequency']).upper().strip()
                        if freq_value in ['MONTHLY', 'QUARTERLY', 'YEARLY', 'ONE_TIME']:
                            payment_frequency = freq_value
                        else:
                            row_auto_gen.append('payment_frequency')
                    else:
                        row_auto_gen.append('payment_frequency')
                    
                    # Generate membership number
                    membership_number = None
                    if 'membership_number' in row and not pd.isna(row['membership_number']):
                        membership_number = str(row['membership_number']).strip()
                        if Member.objects.filter(membership_number=membership_number).exists():
                            membership_number = None
                    
                    if not membership_number:
                        last_member = Member.objects.order_by('-id').first()
                        if last_member and last_member.membership_number:
                            try:
                                last_number = int(last_member.membership_number.split('-')[-1])
                                membership_number = f"NSS-{str(last_number + 1).zfill(3)}"
                            except:
                                membership_number = f"NSS-{str(Member.objects.count() + 1).zfill(3)}"
                        else:
                            membership_number = "NSS-001"
                        row_auto_gen.append('membership_number')
                    
                    # Email
                    email = ''
                    if 'email' in row and not pd.isna(row['email']) and str(row['email']).strip():
                        email = str(row['email']).strip()
                    else:
                        email = f"noemail.{membership_number.lower().replace('-', '')}@placeholder.com"
                        row_auto_gen.append('email')
                    
                    # Address
                    address = ''
                    if 'address' in row and not pd.isna(row['address']) and str(row['address']).strip():
                        address = str(row['address']).strip()
                    else:
                        address = 'Address not provided'
                        row_auto_gen.append('address')
                    
                    # Get optional fields - use None for unique fields if empty
                    father_name = None
                    if 'father_name' in row and not pd.isna(row['father_name']) and str(row['father_name']).strip():
                        father_name = str(row['father_name']).strip()
                    
                    grandfather_name = None
                    if 'grandfather_name' in row and not pd.isna(row['grandfather_name']) and str(row['grandfather_name']).strip():
                        grandfather_name = str(row['grandfather_name']).strip()
                    
                    spouse_name = None
                    if 'spouse_name' in row and not pd.isna(row['spouse_name']) and str(row['spouse_name']).strip():
                        spouse_name = str(row['spouse_name']).strip()
                    
                    # IMPORTANT: Use None instead of empty string for citizenship_number (unique field)
                    citizenship_number = None
                    if 'citizenship_number' in row and not pd.isna(row['citizenship_number']) and str(row['citizenship_number']).strip():
                        citizenship_number = str(row['citizenship_number']).strip()
                    
                    citizenship_issue_district = None
                    if 'citizenship_issue_district' in row and not pd.isna(row['citizenship_issue_district']) and str(row['citizenship_issue_district']).strip():
                        citizenship_issue_district = str(row['citizenship_issue_district']).strip()
                    
                    # Create member
                    member = Member(
                        membership_number=membership_number,
                        name=name,
                        phone=phone,
                        email=email,
                        date_of_birth=date_of_birth,
                        gender=gender,
                        address=address,
                        
                        # Family information (use None for empty)
                        father_name=father_name,
                        grandfather_name=grandfather_name,
                        spouse_name=spouse_name,
                        
                        # Citizenship (use None for unique field)
                        citizenship_number=citizenship_number,
                        citizenship_issue_date=citizenship_issue_date,
                        citizenship_issue_district=citizenship_issue_district,
                        
                        # Membership
                        membership_type=membership_type,
                        payment_frequency=payment_frequency,
                        join_date=join_date,
                        is_active=True
                    )
                    
                    member.save()
                    success_count += 1
                    
                    if row_auto_gen:
                        auto_generated.append({
                            'row': index + 2,
                            'name': name,
                            'fields': row_auto_gen
                        })
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f'Row {index + 2}: {str(e)}')
            
            # Show results
            if success_count > 0:
                messages.success(request, f'✅ Successfully uploaded {success_count} member(s).')
            
            if skipped_count > 0:
                messages.info(request, f'ℹ️ Skipped {skipped_count} empty row(s).')
            
            if auto_generated:
                auto_gen_summary = []
                for item in auto_generated[:10]:
                    fields_str = ', '.join(item['fields'])
                    auto_gen_summary.append(f"Row {item['row']} ({item['name']}): {fields_str}")
                
                auto_gen_message = '<br>'.join(auto_gen_summary)
                if len(auto_generated) > 10:
                    auto_gen_message += f'<br>... and {len(auto_generated) - 10} more rows'
                
                messages.warning(
                    request,
                    f'⚠️ Auto-generated placeholder data for {len(auto_generated)} member(s):<br>{auto_gen_message}<br><br>'
                    f'<strong>Please update these fields later:</strong><br>'
                    f'- Phone numbers starting with "TEMP"<br>'
                    f'- Emails with "@placeholder.com"<br>'
                    f'- Addresses saying "Address not provided"'
                )
            
            if error_count > 0:
                if len(errors) <= 10:
                    error_list = '<br>'.join(errors)
                else:
                    error_list = '<br>'.join(errors[:10]) + f'<br>... and {len(errors) - 10} more'
                
                messages.error(request, f'❌ {error_count} row(s) failed:<br>{error_list}')
            
            return redirect('membership:member_list')
            
        except Exception as e:
            messages.error(request, f'❌ Error processing file: {str(e)}')
            return redirect('membership:bulk_upload_members')
    
    return render(request, 'membership/bulk_upload_members.html')


@login_required
@user_passes_test(is_admin)
def download_template(request):
    """Download Excel template for bulk upload"""
    
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members Template"
    
    # Define headers
    headers = [
        'name', 'phone', 'email', 'date_of_birth', 'gender',
        'address', 'father_name', 'grandfather_name', 'spouse_name',
        'citizenship_number', 'citizenship_issue_date', 'citizenship_issue_district',
        'membership_type', 'join_date'
    ]
    
    # Write headers
    ws.append(headers)
    
    # Add sample data
    sample_data = [
        [
            'John Doe',           # name
            '9841234567',         # phone
            'john@example.com',   # email
            '1990-01-15',         # date_of_birth (YYYY-MM-DD)
            'M',                  # gender (M/F/O)
            'Kathmandu, Nepal',   # address
            'Father Name',        # father_name
            'Grandfather Name',   # grandfather_name
            'Spouse Name',        # spouse_name
            '12-34-567890',       # citizenship_number
            '2010-05-20',         # citizenship_issue_date (YYYY-MM-DD)
            'Kathmandu',          # citizenship_issue_district
            'REGULAR',            # membership_type (REGULAR/LIFETIME)
            '2024-01-01'          # join_date (YYYY-MM-DD)
        ]
    ]
    
    for row_data in sample_data:
        ws.append(row_data)
    
    # Style the header row
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        ["Bulk Upload Instructions"],
        [""],
        ["Required Fields (must be filled):"],
        ["- name: Full name of the member"],
        ["- phone: Contact number"],
        ["- membership_type: Either REGULAR or LIFETIME"],
        [""],
        ["Optional Fields:"],
        ["- email: Email address"],
        ["- date_of_birth: Format YYYY-MM-DD (e.g., 1990-12-31)"],
        ["- gender: M for Male, F for Female, O for Other"],
        ["- address: Full address"],
        ["- father_name, grandfather_name, spouse_name: Family details"],
        ["- citizenship_number, citizenship_issue_date, citizenship_issue_district"],
        ["- join_date: Format YYYY-MM-DD (default: today)"],
        [""],
        ["Date Format:"],
        ["All dates should be in YYYY-MM-DD format"],
        ["Examples: 2024-12-06, 1990-01-15, 2010-05-20"],
        [""],
        ["Membership Type:"],
        ["Use exactly: REGULAR or LIFETIME (case-insensitive)"],
        [""],
        ["Gender:"],
        ["Use: M (Male), F (Female), or O (Other)"],
        [""],
        ["Tips:"],
        ["- Delete the sample row before adding your data"],
        ["- Do not modify the header row"],
        ["- Leave cells empty if data is not available"],
        ["- Phone numbers can include country code"],
    ]
    
    for row in instructions:
        ws_instructions.append(row)
    
    # Adjust column widths
    for ws_temp in [ws, ws_instructions]:
        for column in ws_temp.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_temp.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Create response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=members_upload_template.xlsx'
    
    return response










